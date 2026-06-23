"""Unit Mix calculator (Proforma E1:N28) from rent roll Excel files."""

from __future__ import annotations

from collections import OrderedDict
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

RENT_ROLL_START_ROW = 8
RENT_ROLL_END_ROW = 67

COLUMNS = [
    ("unit_type", "Unit Type (SF)", "text"),
    ("units", "# of Units", "integer"),
    ("occupied", "# of Occupied Units", "integer"),
    ("vacant", "# of Vacant Units", "integer"),
    ("pct", "%", "percent"),
    ("gpr", "GPR", "currency"),
    ("loss_to_lease", "Loss to Lease", "currency"),
    ("in_place", "In Place Rent", "currency"),
    ("vacancy", "Vacancy", "currency"),
    ("net", "Net Rental Income", "currency"),
]


def _is_vacant_or_down(name: Any) -> bool:
    if name is None:
        return False
    return str(name).strip().upper() in ("VACANT", "DOWN")


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _average(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def _read_rent_roll_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.active

    rows: list[dict[str, Any]] = []
    for row_idx in range(RENT_ROLL_START_ROW, RENT_ROLL_END_ROW + 1):
        sq_ft = sheet.cell(row=row_idx, column=3).value  # column C
        if sq_ft is None or str(sq_ft).strip() == "":
            continue

        rows.append(
            {
                "sq_ft": str(sq_ft).strip(),
                "name": sheet.cell(row=row_idx, column=5).value,  # column E
                "market_rent": sheet.cell(row=row_idx, column=6).value,  # column F
                "actual_rent": sheet.cell(row=row_idx, column=7).value,  # column G
            }
        )

    workbook.close()
    return rows


def _unique_unit_types(rows: list[dict[str, Any]]) -> list[str]:
    seen: OrderedDict[str, None] = OrderedDict()
    for row in rows:
        seen.setdefault(row["sq_ft"], None)
    return list(seen.keys())


def calculate_unit_mix(file_bytes: bytes) -> dict[str, Any]:
    rent_rows = _read_rent_roll_rows(file_bytes)
    if not rent_rows:
        raise ValueError("No rent roll data found in rows 8–67 (column C).")

    unit_types = _unique_unit_types(rent_rows)
    data_rows: list[dict[str, float | str]] = []

    for unit_type in unit_types:
        subset = [row for row in rent_rows if row["sq_ft"] == unit_type]
        units = len(subset)
        vacant = sum(1 for row in subset if _is_vacant_or_down(row["name"]))
        occupied = units - vacant

        market_rents = [
            value
            for row in subset
            if (value := _to_float(row["market_rent"])) is not None
        ]
        gpr = _average(market_rents)

        occupied_rows = [
            row for row in subset if not _is_vacant_or_down(row["name"])
        ]
        in_place_values = [
            value
            for row in occupied_rows
            if (value := _to_float(row["actual_rent"])) is not None
        ]
        in_place = _average(in_place_values)

        actual_values = [
            value
            for row in subset
            if (value := _to_float(row["actual_rent"])) is not None
        ]
        net = _average(actual_values)

        data_rows.append(
            {
                "unit_type": unit_type,
                "units": units,
                "occupied": occupied,
                "vacant": vacant,
                "pct": 0.0,
                "gpr": gpr,
                "loss_to_lease": in_place - gpr,
                "in_place": in_place,
                "vacancy": net - in_place,
                "net": net,
            }
        )

    total_units = sum(int(row["units"]) for row in data_rows)
    for row in data_rows:
        row["pct"] = row["units"] / total_units if total_units else 0.0

    total_occupied = sum(int(row["occupied"]) for row in data_rows)
    total_vacant = sum(int(row["vacant"]) for row in data_rows)

    monthly_gpr = sum(row["gpr"] * row["units"] for row in data_rows)
    monthly_loss = sum(row["loss_to_lease"] * row["units"] for row in data_rows)
    monthly_in_place = sum(row["in_place"] * row["units"] for row in data_rows)
    monthly_net = sum(row["net"] * row["units"] for row in data_rows)
    monthly_vacancy = monthly_net - monthly_in_place

    return {
        "title": "Unit Mix",
        "columns": [{"key": key, "label": label, "format": fmt} for key, label, fmt in COLUMNS],
        "rows": data_rows,
        "totals": {
            "label": "Totals",
            "units": total_units,
            "occupied": total_occupied,
            "vacant": total_vacant,
        },
        "monthly": {
            "label": "Monthly",
            "gpr": monthly_gpr,
            "loss_to_lease": monthly_loss,
            "in_place": monthly_in_place,
            "vacancy": monthly_vacancy,
            "net": monthly_net,
        },
        "annually": {
            "label": "Annually",
            "gpr": monthly_gpr * 12,
            "loss_to_lease": monthly_loss * 12,
            "in_place": monthly_in_place * 12,
            "vacancy": monthly_vacancy * 12,
            "net": monthly_net * 12,
        },
    }
